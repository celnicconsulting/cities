# ====================IMPORTS====================
"""MCERT | Cities, Environment, Regions and Transport.

This file is intended to read as a VISUAL SPECIFICATION of the application: the
comments describe what appears on screen, in what order, and what each colour and
marker means, so the interface can be reviewed and rebuilt from the source alone
without running it. Every `render_*` docstring is a layout description, top to
bottom, of the block it draws.

====================================================================
SCREEN LAYOUT
====================================================================

    ┌──────────────┬───────────────────────────────────────────────┐
    │  SIDEBAR     │  H1  MCERT | Cities, Environment, Regions     │
    │              │      and Transport                            │
    │ 🏙️ title      │                                               │
    │ caption      │  ╔═══════════════════════════════════════════╗ │
    │              │  ║ HAZARD BANNER (yellow/black, above tabs)  ║ │
    │ Latest       │  ║ NOT AN OFFICIAL GOVERNMENT PRODUCT        ║ │
    │ published    │  ╚═══════════════════════════════════════════╝ │
    │  · HPI       │                                               │
    │  · Sales     │  ┌ tab bar ──────────────────────────────────┐ │
    │  · Vintage   │  │ 📊 │ 🗺️ │ 💰 │ 📋 │ ⚙️ │ 📓             │ │
    │              │  └───────────────────────────────────────────┘ │
    │ Series       │                                               │
    │ selector     │  active tab body                              │
    │              │                                               │
    │ ℹ️ about      │                                               │
    └──────────────┴───────────────────────────────────────────────┘

Tabs, left to right:

    1  📊 National Overview   real   prices and sales, nationally and by region
    2  🗺️ Price Map           real   district prices as bubbles, animated by year
    3  💰 Area Explorer       real   one district, incl. the value-quartile cut
    4  📋 Data Explorer       real   the detail, filtered, with Excel export
    5  ⚙️ Pipeline            real   provenance, lineage, reconciliation
    6  📓 Build Notes         real   how the platform was built

====================================================================
VISUAL VOCABULARY - the same meaning everywhere
====================================================================

    BRAND green  #1F6F5C   measured price level - the index, median price
    ACCENT red   #E4572E   change and pressure - annual % change
    INK    dark  #12312A   reference marks: the national baseline, tooltips
    blue→red ramp          price level on the map; red is more expensive
    bubble AREA            the sized quantity, on a scale fixed across years
    yellow/black stripes   provenance warning; read before any figure
    ◇  hollow diamond      derived rather than measured
    📋 heading + 📥 Excel  every detail table, button right-justified

**There is no synthetic data in this platform.** The published record is deep
enough that nothing the design wanted needed modelling, so there is no 🔶 marker
and no synthetic banner. The one derived thing is map *position*: districts are
drawn on bundled centroids because MCERT publishes names without coordinates.
The prices themselves are measured.

====================================================================
THE SERIES SELECTOR, AND WHY IT IS NOT A DETAIL
====================================================================

The source publishes two different statistics in one sheet: `Monthly` and
`3-Month rolling`. They are not two periods of one series - they are two
methodologies. Every query in this file fixes STAT_TYPE, and the sidebar makes
the choice explicit rather than silently picking one. The default is 3-Month
rolling, because at the latest period the monthly index covers 41 districts and
the rolling one covers 61: small districts have too few monthly sales to publish.

====================================================================
CODE STRUCTURE
====================================================================

Follows the `snowflake-streamlit-development` template - cached `get_*` data
methods, `render_*` visual methods, a thin `main()`. The one departure is the
session layer: this build runs on Streamlit Community Cloud rather than
Streamlit in Snowflake, so `get_active_session()` is replaced by a read-only
DuckDB connection behind the same function boundary. Every data method keeps its
`df_db_schema` argument, so moving the app into Snowflake means swapping
`run_query` for `session.sql(...).to_pandas()` and nothing else.
"""

import io
import sys
import time
from pathlib import Path

import duckdb
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import pydeck as pdk
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(APP_DIR))
sys.path.insert(0, str(APP_DIR.parent / "scripts"))

from org_banner import Agency, render_provenance_banner  # noqa: E402
from org_build_notes import (TAB_LABEL, load_build_notes,  # noqa: E402
                             notes_fingerprint, notes_path)

st.set_page_config(layout="wide",
                   page_title="MCERT | Cities, Environment, Regions and Transport",
                   page_icon="🏙️")

# ====================SESSION====================
DB_CANDIDATES = [
    APP_DIR.parent / "data" / "cities_public.duckdb",     # public_repo
    APP_DIR.parent / "public" / "cities_public.duckdb",   # working project
    Path("data/cities_public.duckdb"),                    # cwd
]

# Kept for parity with the Snowflake template; unused against DuckDB, where the
# schemas are attached in the same file.
df_db = "PUBLIC"
df_schema_name = "MART"
df_db_schema = f"{df_db}.{df_schema_name}"

BRAND = "#1F6F5C"   # measured price level
ACCENT = "#E4572E"  # change and pressure
INK = "#12312A"     # reference marks and baselines

WELLINGTON = (-41.2865, 174.7762)

AGENCIES = [Agency("Ministry of Housing and Urban Development",
                   sub="MCERT House Price Index")]

BUILD_NOTES = notes_path(__file__)


def _extract_fingerprint():
    """Path, size and modified time of the extract, used as a cache key.

    Streamlit Community Cloud hot-reloads on a push: it pulls the new files and
    re-runs the script, but does NOT clear `cache_resource`. A connection opened
    before the pull therefore keeps reading the replaced file, and a query
    written against a newly added column fails to bind against data sitting
    correctly on disk a few bytes away. Making the fingerprint an argument turns
    a data refresh into a cache miss, so the connection reopens by itself.
    """
    for path in DB_CANDIDATES:
        if path.exists():
            s = path.stat()
            return (str(path), s.st_size, int(s.st_mtime))
    return None


@st.cache_resource(show_spinner=False)
def _open_connection(fingerprint):
    if not fingerprint:
        st.error("No data extract found. Expected `data/cities_public.duckdb`.")
        st.stop()
    return duckdb.connect(fingerprint[0], read_only=True)


def get_connection():
    return _open_connection(_extract_fingerprint())


@st.cache_data(show_spinner=False)
def _run_query_cached(sql: str, fingerprint) -> pd.DataFrame:
    return get_connection().execute(sql).fetchdf()


def run_query(sql: str) -> pd.DataFrame:
    """Every read goes through here, so porting to Snowflake is one function."""
    return _run_query_cached(sql, _extract_fingerprint())


# ====================DATA====================
# One method per visual element. Each names the element it feeds, and every one
# fixes STAT_TYPE - the two published series must never interleave.
#
# **These are deliberately not decorated with `@st.cache_data`.** Caching happens
# one level down, in `_run_query_cached`, which is keyed on the SQL *and* the
# extract fingerprint. Decorating these instead keys the cache on the caller's
# arguments only, so the fingerprint is never consulted: the outer cache returns
# a frame from the previous extract without ever calling the inner one, and a
# data refresh appears to do nothing. The result is the whole point of the
# fingerprint, silently defeated one layer up.
def get_latest(df_db_schema):
    """Sidebar: latest periods, the vintage, and coverage counts."""
    return run_query("SELECT * FROM MART.M_LATEST")


def get_stat_types(df_db_schema):
    """Sidebar selector: the statistical series the source publishes."""
    return run_query("SELECT DISTINCT STAT_TYPE FROM MART.M_HPI_NATIONAL "
                     "ORDER BY STAT_TYPE DESC")["STAT_TYPE"].tolist()


def get_national(df_db_schema, stat_type):
    """T1.C1 metrics and T1.C2/C3 lines: the national index and sales."""
    return run_query(
        "SELECT PERIOD, PERIOD_DATE, HPI, ANNUAL_CHANGE_PCT, SALES, MEDIAN_PRICE "
        f"FROM MART.M_HPI_NATIONAL WHERE STAT_TYPE = '{stat_type}' ORDER BY PERIOD")


def get_regions(df_db_schema, stat_type):
    """T1.C2 series and T1.C4 bars: the index by region."""
    return run_query(
        "SELECT PERIOD, PERIOD_DATE, REGION, HPI, ANNUAL_CHANGE_PCT "
        f"FROM MART.M_HPI_REGION WHERE STAT_TYPE = '{stat_type}' ORDER BY PERIOD")


def get_ta_latest(df_db_schema, stat_type):
    """T2.C2 map and T2.C3 table: every district at the latest published period."""
    return run_query(
        "SELECT TA_NAME, REGION, LAT, LON, HPI, ANNUAL_CHANGE_PCT, SALES_USED, "
        f"MEDIAN_PRICE, PERIOD FROM MART.M_HPI_TA_LATEST WHERE STAT_TYPE = '{stat_type}' "
        "AND HPI IS NOT NULL ORDER BY HPI DESC")


def get_map_years(df_db_schema, stat_type):
    """T2: every district in every year, for the animated map.

    The WHOLE table for this statistical series is fetched once and filtered by
    year in memory. At roughly 3,500 rows per series that costs nothing, and it
    is what keeps the play button smooth: a database round trip inside the
    animation loop would stutter at one frame every few seconds.
    """
    return run_query(
        "SELECT YEAR, TA_NAME, REGION, LAT, LON, PERIOD_USED, MEDIAN_PRICE, "
        "SALES_YEAR, EST_VALUE_YEAR, HPI, ANNUAL_CHANGE_PCT "
        f"FROM MART.M_MAP_TA_YEAR WHERE STAT_TYPE = '{stat_type}' "
        "ORDER BY YEAR, TA_NAME")


def get_area_options(df_db_schema, area_type):
    """T3.C1 and T4.C1 selectors."""
    return run_query(
        f"SELECT DISTINCT AREA FROM MART.M_HPI_AREA WHERE AREA_TYPE = '{area_type}' "
        "ORDER BY AREA")["AREA"].tolist()


def get_area_series(df_db_schema, area_type, areas, stat_type):
    """T3.C2: the index for the selected areas."""
    if not areas:
        return pd.DataFrame()
    lst = ", ".join("'" + a.replace("'", "''") + "'" for a in areas)
    return run_query(
        "SELECT PERIOD, PERIOD_DATE, AREA, HPI, ANNUAL_CHANGE_PCT FROM MART.M_HPI_AREA "
        f"WHERE AREA_TYPE = '{area_type}' AND AREA IN ({lst}) "
        f"AND STAT_TYPE = '{stat_type}' ORDER BY PERIOD")


def get_quartiles(df_db_schema, area_type, area, stat_type):
    """T3.C3: the value-quartile cut - the split no other NZ index publishes."""
    a = area.replace("'", "''")
    return run_query(
        "SELECT PERIOD, PERIOD_DATE, VALUE_QUARTILE, HPI FROM MART.M_HPI_QUARTILE "
        f"WHERE AREA_TYPE = '{area_type}' AND AREA = '{a}' "
        f"AND STAT_TYPE = '{stat_type}' ORDER BY PERIOD")


def get_sales_series(df_db_schema, area_type, area, stat_type):
    """T3.C4: lower quartile, median and upper quartile sale price."""
    a = area.replace("'", "''")
    return run_query(
        "SELECT PERIOD, PERIOD_DATE, SALES, LOWER_QUARTILE, MEDIAN_PRICE, "
        f"UPPER_QUARTILE FROM MART.M_SALES_AREA WHERE AREA_TYPE = '{area_type}' "
        f"AND AREA = '{a}' AND STAT_TYPE = '{stat_type}' ORDER BY PERIOD")


def get_detail(df_db_schema, area_type, areas, stat_type, quartile, p_from, p_to):
    """T4.C2: the filtered detail fact behind the Excel export."""
    if not areas:
        return pd.DataFrame()
    lst = ", ".join("'" + a.replace("'", "''") + "'" for a in areas)
    q = "" if quartile == "All quartiles" else f"AND VALUE_QUARTILE = '{quartile}' "
    return run_query(
        "SELECT PERIOD, AREA_TYPE, AREA, VALUE_QUARTILE, HPI, ANNUAL_CHANGE_PCT, "
        f"SALES_USED FROM MART.M_HPI_QUARTILE WHERE AREA_TYPE = '{area_type}' "
        f"AND AREA IN ({lst}) AND STAT_TYPE = '{stat_type}' {q}"
        f"AND PERIOD BETWEEN '{p_from}' AND '{p_to}' ORDER BY PERIOD, AREA")


def get_validation(df_db_schema):
    """T5.C5: the reconciliation, shown rather than asserted."""
    return run_query("SELECT CHECK_NAME, EXPECTED, ACTUAL, PASSED, NOTE "
                     "FROM MART.VALIDATION_RESULTS")


def get_sources(df_db_schema):
    """T5.C2: the source register, including what was NOT ingested and why."""
    return run_query("SELECT DATASET_ID, ORG, SOURCE_NAME, INGESTED, CADENCE, "
                     "LICENCE, URL, NOTES FROM MART.META_SOURCE_REGISTER "
                     "ORDER BY DATASET_ID")


def get_downloads(df_db_schema):
    """T5.C3: every source file with its size, checksum and download date."""
    return run_query("SELECT org, dataset_id, file_name, bytes, md5, fetched_at, "
                     "status, file_url FROM MART.META_DOWNLOAD ORDER BY dataset_id, file_name")


def get_raw_catalog(df_db_schema):
    """T5.C4: one row per RAW table, with period span and inferred cadence."""
    return run_query("SELECT SCHEMA_NAME, TABLE_NAME, LAYOUT, N_ROWS, N_COLS, "
                     "PERIOD_MIN, PERIOD_MAX, CADENCE FROM MART.META_RAW_CATALOG "
                     "ORDER BY CAST(N_ROWS AS BIGINT) DESC")


def get_aliases(df_db_schema):
    """T5: the area labels that were remapped, so the substitution is auditable."""
    return run_query("SELECT AREA_TYPE, AREA_RAW, AREA_CANONICAL FROM MART.STG_AREA_ALIAS "
                     "WHERE REMAPPED = 'YES' ORDER BY AREA_TYPE, AREA_RAW")


@st.cache_data(show_spinner=False)
def get_build_notes(fingerprint):
    """T6: the Build Notes markdown, loaded from disk rather than embedded."""
    return load_build_notes(BUILD_NOTES, fingerprint)


# ====================STATIC_METHODS====================
def build_styled_excel(df: pd.DataFrame, title: str) -> bytes:
    """A branded workbook: title bar, styled header, frozen panes, auto-filter."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ncols = max(1, len(df.columns))
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="FFFFFF")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=BRAND.lstrip("#"))
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=2, column=j, value=str(col))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INK.lstrip("#"))
    for i, row in enumerate(df.itertuples(index=False), start=3):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=(None if pd.isna(v) else v))
        if i % 2 == 0:
            for j in range(1, ncols + 1):
                ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor="F2F7F5")
    for j, col in enumerate(df.columns, start=1):
        width = max(10, min(38, int(df[col].astype(str).str.len().max() or 10) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{len(df) + 2}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def detail_table(df: pd.DataFrame, title: str, key: str):
    """📋 heading with the 📥 Excel button right-justified on the same row."""
    head, btn = st.columns([3, 1])
    head.markdown(f"**📋 {title}**")
    if not df.empty:
        btn.download_button("📥 Excel", build_styled_excel(df, title),
                            file_name=f"{key}.xlsx", key=f"dl_{key}",
                            mime="application/vnd.openxmlformats-officedocument."
                                 "spreadsheetml.sheet",
                            use_container_width=True)
    st.dataframe(df, use_container_width=True, hide_index=True)


# ====================SIDEBAR====================
def render_sidebar():
    """Title, the latest published periods and vintage, the series selector, about.

    The series selector is here rather than on a tab because it changes every
    figure in the application - it chooses which of the source's two statistics
    is being read.
    """
    with st.sidebar:
        st.markdown("### 🏙️ MCERT")
        st.caption("Cities, Environment, Regions and Transport — built from "
                   "public releases by the Ministry of Housing and Urban "
                   "Development, whose functions moved to MCERT on 1 July 2026.")

        latest = get_latest(df_db_schema).iloc[0]
        st.markdown("**Latest published**")
        st.markdown(
            f"- House price index · `{latest['LATEST_HPI_PERIOD']}`\n"
            f"- Sales statistics · `{latest['LATEST_SALES_PERIOD']}`\n"
            f"- Series begins · `{latest['EARLIEST_PERIOD']}`\n"
            f"- Districts · `{int(latest['N_TA'])}`"
        )
        st.markdown("**Workbook vintage**")
        st.info(f"`{latest['VINTAGE']}` — the whole series is recalculated each "
                "month, so these figures are this vintage's estimate of all "
                "56 years, not a fixed history.", icon="🗓️")

        st.divider()
        stat_types = get_stat_types(df_db_schema)
        default = stat_types.index("3-Month rolling") if "3-Month rolling" in stat_types else 0
        stat_type = st.radio("Statistical series", stat_types, index=default,
                             help="The source publishes two different statistics "
                                  "in one sheet. They are never mixed. At the "
                                  "latest period the monthly index covers 41 "
                                  "districts; the 3-month rolling one covers 61.")

        st.divider()
        st.markdown("**Provenance**")
        st.markdown("- All figures are **measured**, from published releases.\n"
                    "- ◇ Map *position* is derived from district centroids.\n"
                    "- No synthetic data in this platform.")
        with st.expander("ℹ️ About"):
            st.markdown(
                "Built by Celnic Consulting from public data. The charted "
                "figures are used under **CC BY 4.0**, verified against the "
                "publisher's copyright page on 30 Aug 2026. Not an official "
                "government product.\n\n"
                "**1 July 2026** — the Ministry for the Environment, the "
                "Ministry of Housing and Urban Development, the Ministry of "
                "Transport and the local government functions of the Department "
                "of Internal Affairs merged into MCERT."
            )
    return stat_type


# ====================VISUALISATION====================
def render_tab_national(stat_type):
    """Four headline metrics, then the national index, sales, and regional change.

    Top row: index level, annual change, monthly sales, median price, each with
    a delta against the previous published period. Then a multi-region line with
    the national series drawn in INK as a baseline, a national sales line, and a
    bar of annual change by region sorted descending.
    """
    nat = get_national(df_db_schema, stat_type)
    nat = nat[nat["HPI"].notna()]
    if nat.empty:
        st.warning("No national series for this statistical series.")
        return
    cur, prev = nat.iloc[-1], nat.iloc[-2]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("House price index", f"{cur['HPI']:,.0f}",
              f"{cur['HPI'] - prev['HPI']:+,.0f} on previous month")
    # Annual change is published nearly two years behind the index itself, so
    # the headline takes the latest period it actually exists for and says so,
    # rather than showing an em dash on a metric that is simply older.
    ac = nat[nat["ANNUAL_CHANGE_PCT"].notna()]
    if ac.empty:
        c2.metric("Annual change", "—")
    else:
        acr = ac.iloc[-1]
        c2.metric("Annual change", f"{acr['ANNUAL_CHANGE_PCT']:.1f}%",
                  f"as at {acr['PERIOD']}", delta_color="off")
    c3.metric("Residential sales",
              "—" if pd.isna(cur["SALES"]) else f"{cur['SALES']:,.0f}")
    c4.metric("Median sale price",
              "—" if pd.isna(cur["MEDIAN_PRICE"]) else f"${cur['MEDIAN_PRICE']:,.0f}")
    st.caption(f"Latest published period: **{cur['PERIOD']}** · series: **{stat_type}**")

    st.divider()
    regions = get_regions(df_db_schema, stat_type)
    opts = sorted(regions["REGION"].dropna().unique().tolist())
    default = [r for r in ["Auckland", "Wellington", "Canterbury", "Otago"] if r in opts]
    picked = st.multiselect("Regions", opts, default=default or opts[:4])

    years = sorted(nat["PERIOD"].str[:4].unique())
    y_from, y_to = st.select_slider("Period range", options=years,
                                    value=(years[max(0, len(years) - 26)], years[-1]))

    sub = regions[regions["REGION"].isin(picked)
                  & regions["PERIOD"].str[:4].between(y_from, y_to)]
    nsub = nat[nat["PERIOD"].str[:4].between(y_from, y_to)]

    fig = px.line(sub, x="PERIOD_DATE", y="HPI", color="REGION",
                  labels={"PERIOD_DATE": "", "HPI": "House price index"})
    fig.add_trace(go.Scatter(x=nsub["PERIOD_DATE"], y=nsub["HPI"], name="National",
                             line=dict(color=INK, width=3, dash="dot")))
    fig.update_layout(height=430, legend_title_text="")
    st.plotly_chart(fig, use_container_width=True)

    left, right = st.columns(2)
    f2 = px.area(nsub, x="PERIOD_DATE", y="SALES",
                 labels={"PERIOD_DATE": "", "SALES": "Residential sales"},
                 color_discrete_sequence=[BRAND])
    f2.update_layout(height=330, title="National residential sales")
    left.plotly_chart(f2, use_container_width=True)

    latest_p = sub["PERIOD"].max() if not sub.empty else None
    bar = regions[(regions["PERIOD"] == latest_p) & regions["ANNUAL_CHANGE_PCT"].notna()]
    bar = bar.sort_values("ANNUAL_CHANGE_PCT", ascending=False)
    f3 = px.bar(bar, x="REGION", y="ANNUAL_CHANGE_PCT",
                labels={"REGION": "", "ANNUAL_CHANGE_PCT": "Annual change %"},
                color_discrete_sequence=[ACCENT])
    f3.update_layout(height=330, title=f"Annual change by region · {latest_p}")
    right.plotly_chart(f3, use_container_width=True)


SIZE_MEASURES = {
    "Median sale price": ("MEDIAN_PRICE", "the district's own median sale price"),
    "Estimated total sales value": ("EST_VALUE_YEAR",
                                    "median price times sales that year - an estimate"),
}


def render_tab_map(stat_type):
    """Bubbles over district centroids, sized by price, animated through the years.

    Top row: what to size by, seconds between frames, max bubble size, play/pause.
    Then a year slider, the map, and the detail table for the year on screen.

    **Bubble area is proportional to the value, on a scale fixed across all
    years.** Two decisions worth stating, because both change what the map means:

    * Area, not radius, carries the value. Radius-proportional circles overstate
      differences by squaring them - a district at twice the price would look
      four times as big.
    * The scale is global, not per-year. Normalising each year to its own maximum
      would make the biggest bubble identical in 1980 and 2026, and the animation
      would show nothing about prices rising. On a fixed scale the whole map
      inflates over time, which is the actual story.

    Bubbles are translucent and allowed to overlap: neighbouring districts
    genuinely do overlap at these radii, and preventing it would mean shrinking
    the bubbles or moving them off their centroids.

    Playback is driven from the browser, so it pauses while the tab is hidden and
    resumes when it is shown again. That is Streamlit's behaviour, not a fault,
    and it is the right one - nothing should burn server time animating a map
    nobody is looking at.
    """
    df = get_map_years(df_db_schema, stat_type)
    if df.empty:
        st.warning("No district series for this statistical series.")
        return

    years = sorted(int(y) for y in df["YEAR"].unique())
    if "map_year" not in st.session_state or st.session_state.map_year not in years:
        st.session_state.map_year = years[-1]
    st.session_state.setdefault("map_playing", False)

    c1, c2, c3, c4 = st.columns([3, 1.3, 1.2, 1.3])
    size_label = c1.radio(
        "Size bubbles by", list(SIZE_MEASURES), horizontal=True,
        help="Median sale price is the district's own price level - highest in "
             "Queenstown-Lakes. Estimated total sales value is that price times "
             "the number of sales, so it reflects market size, which is what puts "
             "Auckland far ahead of everywhere else.")
    secs = c2.number_input("Seconds per year", min_value=1, max_value=30,
                           value=5, step=1)
    max_km = c3.number_input("Max bubble (km)", min_value=10, max_value=120,
                             value=45, step=5)

    if st.session_state.map_playing:
        if c4.button("Pause", use_container_width=True, icon=":material/pause:"):
            st.session_state.map_playing = False
            st.rerun()
    else:
        if c4.button("Play", use_container_width=True,
                     icon=":material/play_arrow:"):
            # Starting on the last year would play nothing, so wrap to the start.
            if st.session_state.map_year >= years[-1]:
                st.session_state.map_year = years[0]
            st.session_state.map_playing = True
            st.session_state.map_last_tick = 0.0   # first frame draws at once
            st.rerun()

    col, blurb = SIZE_MEASURES[size_label]
    vmax = float(df[col].max())   # fixed across every year - see the docstring

    @st.fragment(run_every=(f"{int(secs)}s" if st.session_state.map_playing else None))
    def _frame():
        # Advance on elapsed time, not merely on being re-entered. The fragment
        # reruns for two reasons - the timer, and the year slider inside it
        # re-rendering - so advancing unconditionally moves two years per tick
        # and the configured seconds mean nothing. Gating on the clock makes the
        # setting real whatever triggered the rerun.
        if st.session_state.map_playing:
            now = time.monotonic()
            if now - st.session_state.get("map_last_tick", 0.0) >= int(secs) * 0.9:
                st.session_state.map_last_tick = now
                i = years.index(st.session_state.map_year)
                if i + 1 < len(years):
                    st.session_state.map_year = years[i + 1]
                else:
                    st.session_state.map_playing = False   # stop at the last year

        st.select_slider("Year", options=years, key="map_year")
        year = st.session_state.map_year
        d = df[(df["YEAR"] == year) & df[col].notna()].copy()
        if d.empty:
            st.info(f"No district prices published for {year}.")
            return

        d["RADIUS"] = ((max_km * 1000) * (d[col] / vmax).clip(lower=0) ** 0.5)
        # Total sales value spans roughly 300x across districts, so an
        # area-proportional radius puts the smallest at about a kilometre -
        # invisible at national zoom. The floor keeps every published
        # district on the map; above it the encoding stays proportional.
        d["RADIUS"] = d["RADIUS"].clip(lower=6000)

        # Colour on the same quantity, so size and colour can never disagree.
        lo, hi = d[col].min(), d[col].max()
        tint = (d[col] - lo) / ((hi - lo) or 1)
        d["_r"] = (40 + 200 * tint).astype(int)
        d["_g"] = (110 - 40 * tint).astype(int).clip(lower=0)
        d["_b"] = (200 - 160 * tint).astype(int).clip(lower=0)
        d["PRICE_LABEL"] = d["MEDIAN_PRICE"].map(lambda v: f"${v:,.0f}")
        d["SALES_LABEL"] = d["SALES_YEAR"].map(
            lambda v: "n/a" if pd.isna(v) else f"{v:,.0f}")
        d["VALUE_LABEL"] = d["EST_VALUE_YEAR"].map(
            lambda v: "n/a" if pd.isna(v) else f"${v / 1e9:,.2f}bn")

        layer = pdk.Layer(
            "ScatterplotLayer", data=d, get_position=["LON", "LAT"],
            get_fill_color=["_r", "_g", "_b", 120],       # translucent: overlaps read
            get_radius="RADIUS", pickable=True, stroked=True,
            get_line_color=[255, 255, 255, 160], line_width_min_pixels=1,
        )
        st.pydeck_chart(pdk.Deck(
            layers=[layer],
            initial_view_state=pdk.ViewState(latitude=-41.0, longitude=173.0,
                                             zoom=4.4),
            map_provider="carto", map_style="light",
            tooltip={"text": "{TA_NAME}\n{REGION}\n"
                             "Median price: {PRICE_LABEL}\n"
                             "Sales: {SALES_LABEL}\n"
                             "Est. total value: {VALUE_LABEL}"},
        ))

        top = d.nlargest(1, col).iloc[0]
        if year == years[-1]:
            st.caption(
                f":orange[**{year} is a part year.**] The series ends at "
                f"{st.session_state.get('_latest_period', top['PERIOD_USED'])}, so "
                f"sales and total value for {year} cover only the months published "
                f"so far and are not comparable with a full year. Median price is "
                f"unaffected - it is a level, not a total.")
        st.caption(
            f"**{year}** - sized by {size_label.lower()} ({blurb}). Largest: "
            f"**{top['TA_NAME']}**. {len(d)} districts published, series "
            f"**{stat_type}**. Bubble **area** is proportional to the value on a "
            f"scale fixed across {years[0]}-{years[-1]}, so the whole map grows as "
            f"prices rise. Each year's price is that district's last published "
            f"month ({top['PERIOD_USED']} for {top['TA_NAME']}), never an average "
            f"of the twelve. Estimated total value uses the median as a stand-in "
            f"for the mean, so it is an estimate. Position is derived: districts "
            f"sit on bundled centroids, not boundaries, and bubble size is a "
            f"quantity, not an extent."
        )

        detail_table(
            d[["TA_NAME", "REGION", "YEAR", "PERIOD_USED", "MEDIAN_PRICE",
               "SALES_YEAR", "EST_VALUE_YEAR", "HPI", "ANNUAL_CHANGE_PCT"]]
            .sort_values(col, ascending=False),
            f"Districts in {year} - {stat_type}", "map_year_detail")

    _frame()


def render_tab_area(stat_type):
    """One area at a time: the index, the value-quartile cut, and sale prices.

    The quartile chart is the centrepiece - the source cuts its index by property
    value quartile, which no other New Zealand house price index publishes.
    """
    area_type = st.radio("Area type",
                         ["Territorial Authority", "Region", "Auckland Local Board"],
                         horizontal=True)
    opts = get_area_options(df_db_schema, area_type)
    default = [o for o in ["Wellington City", "Auckland", "Christchurch City"] if o in opts]
    picked = st.multiselect("Compare areas", opts, default=default or opts[:3])

    series = get_area_series(df_db_schema, area_type, picked, stat_type)
    if not series.empty:
        fig = px.line(series, x="PERIOD_DATE", y="HPI", color="AREA",
                      labels={"PERIOD_DATE": "", "HPI": "House price index"})
        fig.update_layout(height=400, legend_title_text="")
        st.plotly_chart(fig, use_container_width=True)

    st.divider()
    focus = st.selectbox("Value-quartile detail for", picked or opts[:1])
    if not focus:
        return
    left, right = st.columns(2)

    q = get_quartiles(df_db_schema, area_type, focus, stat_type)
    q = q[q["HPI"].notna()]
    if not q.empty:
        fq = px.line(q, x="PERIOD_DATE", y="HPI", color="VALUE_QUARTILE",
                     labels={"PERIOD_DATE": "", "HPI": "Index"},
                     category_orders={"VALUE_QUARTILE": ["All", "1", "2", "3", "4"]})
        fq.update_layout(height=380, title=f"{focus} · index by value quartile",
                         legend_title_text="Quartile")
        left.plotly_chart(fq, use_container_width=True)

    s = get_sales_series(df_db_schema, area_type, focus, stat_type)
    s = s[s["MEDIAN_PRICE"].notna()]
    if not s.empty:
        fs = go.Figure()
        for col, name, colour, dash in [
            ("UPPER_QUARTILE", "Upper quartile", ACCENT, "dot"),
            ("MEDIAN_PRICE", "Median", BRAND, None),
            ("LOWER_QUARTILE", "Lower quartile", INK, "dot"),
        ]:
            fs.add_trace(go.Scatter(x=s["PERIOD_DATE"], y=s[col], name=name,
                                    line=dict(color=colour, dash=dash)))
        fs.update_layout(height=380, title=f"{focus} · sale price",
                         yaxis_title="NZD", legend_title_text="")
        right.plotly_chart(fs, use_container_width=True)


def render_tab_detail(stat_type):
    """The filtered fact, with the styled Excel export beside its title."""
    c1, c2, c3 = st.columns(3)
    area_type = c1.selectbox("Area type",
                             ["Territorial Authority", "Region",
                              "Auckland Local Board", "National"])
    opts = get_area_options(df_db_schema, area_type)
    areas = c2.multiselect("Areas", opts, default=opts[:3])
    quartile = c3.selectbox("Value quartile",
                            ["All quartiles", "All", "1", "2", "3", "4"])

    periods = get_national(df_db_schema, stat_type)["PERIOD"].tolist()
    if not periods:
        st.warning("No periods for this series.")
        return
    p_from, p_to = st.select_slider(
        "Period range", options=periods,
        value=(periods[max(0, len(periods) - 60)], periods[-1]))

    df = get_detail(df_db_schema, area_type, areas, stat_type, quartile, p_from, p_to)
    st.caption(f"{len(df):,} rows · series **{stat_type}**")
    detail_table(df, f"House price index detail · {p_from} to {p_to}", "detail")


def render_tab_pipeline():
    """Provenance: counts, the source register incl. exclusions, and reconciliation.

    The pass/fail state is computed from the stored results, so this tab cannot
    claim the checks pass while they are failing.
    """
    val = get_validation(df_db_schema)
    dl = get_downloads(df_db_schema)
    cat = get_raw_catalog(df_db_schema)
    passed = int(val["PASSED"].sum()) if not val.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Source files downloaded", f"{(dl['status'] != 'failed').sum():,}")
    c2.metric("RAW tables", f"{len(cat):,}")
    c3.metric("RAW rows", f"{cat['N_ROWS'].astype(int).sum():,}")
    c4.metric("Checks passing", f"{passed}/{len(val)}")

    if len(val) and passed == len(val):
        st.success(f"All {len(val)} reconciliation checks pass.", icon="✅")
    else:
        st.error(f"{len(val) - passed} of {len(val)} checks are failing.", icon="⚠️")

    st.divider()
    st.markdown("**Source register** — including sources deliberately not "
                "ingested, with the reason. A source that quietly did not appear "
                "would be indistinguishable from one that was never considered.")
    src = get_sources(df_db_schema)
    st.dataframe(src, use_container_width=True, hide_index=True,
                 column_config={"NOTES": st.column_config.TextColumn(width="large")})

    st.divider()
    detail_table(val, "Reconciliation checks", "validation")

    st.divider()
    alias = get_aliases(df_db_schema)
    st.markdown("**Area labels remapped** — MCERT publishes district names with "
                "te reo macrons; the bundled reference geography does not. "
                "Joining on the raw label drops these districts silently.")
    st.dataframe(alias, use_container_width=True, hide_index=True)

    st.divider()
    detail_table(cat, "RAW table catalog", "raw_catalog")
    st.divider()
    detail_table(dl, "Download manifest", "downloads")


def render_tab_build_notes():
    """Full-width rendering of the build notes markdown, loaded from disk."""
    st.markdown(get_build_notes(notes_fingerprint(BUILD_NOTES)))


def render_attribution():
    """Collapsed attribution line below the tabs, on every page.

    The sidebar About expander says who built this; this says whose data it is
    and where the full provenance manifest lives. It sits at the bottom of
    `main` so it appears under whichever tab is open, rather than only under
    the one tab a reader might never select.
    """
    with st.expander("Data sources & attribution"):
        st.markdown(
            "**Charted figures** — Property and Sales Statistics (incl. the "
            "MCERT House Price Index), © Te Tūāpapa Kura Kāinga / MHUD, now "
            "MCERT, used under **CC BY 4.0** "
            "([verified at source](https://www.hud.govt.nz/about-us/copyright-and-disclaimer), "
            "30 Aug 2026 — an agency-wide statement, not a dataset-specific "
            "one).\n\n"
            "**Pipeline metadata** — derived from the MHUD data.govt.nz "
            "catalogue, which is **not uniformly licensed**: 23 of 27 packages "
            "are CC BY 3.0 NZ and 4 state no licence. None of its data is "
            "republished here.\n\n"
            "Data is modified (downloaded, staged, transformed) and not "
            "synthetic; demonstration of method, not published statistics. "
            "Full provenance: "
            "[ATTRIBUTION.md](https://github.com/celnicconsulting/cities/blob/main/ATTRIBUTION.md)."
        )


# ====================MAIN====================
def main():
    st.title("MCERT | Cities, Environment, Regions and Transport")
    render_provenance_banner(AGENCIES)
    stat_type = render_sidebar()

    tabs = st.tabs(["📊 National Overview", "🗺️ Price Map", "💰 Area Explorer",
                    "📋 Data Explorer", "⚙️ Pipeline", TAB_LABEL])
    with tabs[0]:
        render_tab_national(stat_type)
    with tabs[1]:
        render_tab_map(stat_type)
    with tabs[2]:
        render_tab_area(stat_type)
    with tabs[3]:
        render_tab_detail(stat_type)
    with tabs[4]:
        render_tab_pipeline()
    with tabs[5]:
        render_tab_build_notes()

    render_attribution()


if __name__ == "__main__":
    main()
